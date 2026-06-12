"""ReActスタイルのExcel編集エージェント。

LLMが提案するopenpyxlコードを「永続サンドボックス子プロセス」で1ステップずつ実行し、
print()や例外を観測としてLLMへ戻して次の手を促す。複数ステップ・自己修正に対応する。

権限方針:
    元データはユーザーPC側にあり、サーバーが触るのはアップロードされた複製のため、
    ファイル破損リスクは低い。一方でサーバー自体のRCEは防ぐ必要があるため、
    - import はホワイトリスト(_SAFE_IMPORT_ROOTS)のみ許可（datetime/re/openpyxl.styles等）
    - os/sys/subprocess/socket/open/eval/exec 等はビルトインから除去
    という形で「Excel編集に必要な範囲だけ」開放している。
"""

import ast
import io
import multiprocessing as mp
import re
import traceback
from contextlib import redirect_stdout
from pathlib import Path
from typing import Any

# import 可能な標準ライブラリのルート（サブモジュールも許可）。
# I/O・システムアクセスを伴うモジュールは含めない。
_SAFE_IMPORT_ROOTS = {
    "datetime",
    "re",
    "math",
    "decimal",
    "fractions",
    "statistics",
    "calendar",
    "collections",
    "string",
    "itertools",
    "functools",
    "json",
    "copy",
    "textwrap",
    "unicodedata",
}

# openpyxl は「丸ごと」は許可しない（`from openpyxl import load_workbook` で
# サーバー上の任意ブックを読み出し print() 経由で内容を漏洩できてしまうため）。
# ファイルI/Oを伴わない書式・ユーティリティ系サブモジュールだけを許可する。
_SAFE_OPENPYXL_MODULES = {
    "openpyxl.styles",
    "openpyxl.utils",
    "openpyxl.utils.cell",
    "openpyxl.comments",
    "openpyxl.formatting",
    "openpyxl.formatting.rule",
    "openpyxl.worksheet.table",
    "openpyxl.chart",
}

# 安全なモジュール経由でも到達させてはいけない危険モジュール／属性名。
# precheck で ast.Name / ast.Attribute の両方として遮断する
# （例: `import random; random.os.system(...)` のような属性チェーン脱獄を防ぐ）。
_DANGEROUS_NAMES = {
    "os",
    "sys",
    "subprocess",
    "socket",
    "shutil",
    "ctypes",
    "importlib",
    "posix",
    "nt",
    "builtins",
    "platform",
    "multiprocessing",
    "threading",
    "signal",
    "pickle",
    "marshal",
    "inspect",
    "pty",
    "system",
    "popen",
    "spawn",
}

# サンドボックスで公開するビルトインの「明示的許可リスト」。
# dir(builtins) からの除外方式だと site が注入する license/copyright/help 等が
# 残り、`license._Printer__filenames` 経由で任意ファイルを開示できてしまうため、
# 必要なものだけを列挙する方式にする（type/object/super/getattr 等は意図的に除外）。
_ALLOWED_BUILTINS = {
    "abs", "all", "any", "ascii", "bin", "bool", "bytearray", "bytes",
    "callable", "chr", "dict", "divmod", "enumerate", "filter", "float",
    "format", "frozenset", "hasattr", "hash", "hex", "int", "isinstance",
    "issubclass", "iter", "len", "list", "map", "max", "min", "next", "oct",
    "ord", "pow", "print", "range", "repr", "reversed", "round", "set",
    "slice", "sorted", "str", "sum", "tuple", "zip",
}
# try/except で使う例外クラス
_ALLOWED_EXCEPTIONS = {
    "BaseException", "Exception", "ArithmeticError", "AssertionError",
    "AttributeError", "FloatingPointError", "IndexError", "KeyError",
    "LookupError", "NameError", "NotImplementedError", "OverflowError",
    "RuntimeError", "StopIteration", "TypeError", "ValueError",
    "ZeroDivisionError",
}

# 事前ASTチェックで弾く名前参照（ビルトイン除去と二重防御）。
# 危険モジュール名(_DANGEROUS_NAMES)も含め、ast.Name / ast.Attribute の両方で遮断する。
_FORBIDDEN_NAMES = {
    "eval",
    "exec",
    "compile",
    "open",
    "input",
    "breakpoint",
    "getattr",
    "setattr",
    "delattr",
    "globals",
    "locals",
    "vars",
    "__import__",
} | _DANGEROUS_NAMES

MAX_OBSERVATION_CHARS = 1500


def _is_allowed_import(name: str) -> bool:
    root = name.split(".", 1)[0]
    if root in _SAFE_IMPORT_ROOTS:
        return True
    if name in _SAFE_OPENPYXL_MODULES:
        return True
    return any(name.startswith(f"{mod}.") for mod in _SAFE_OPENPYXL_MODULES)


def _safe_import(name, globals=None, locals=None, fromlist=(), level=0):
    import builtins as _b

    if level != 0 or not _is_allowed_import(name):
        raise ImportError(f"このサンドボックスでは '{name}' の import は許可されていません")
    # `import openpyxl.utils` は namespace に親パッケージ openpyxl を束縛し、
    # openpyxl.load_workbook 等のファイルI/Oを露出させてしまう。openpyxl は
    # 親を束縛しない from-import 形式（fromlist あり）のみ許可する。
    if name.split(".", 1)[0] == "openpyxl" and not fromlist:
        raise ImportError(
            "openpyxl は `from openpyxl.styles import ...` の形式でのみ利用できます"
        )
    # fromlist 経由のエイリアス脱獄を遮断する。
    # 例: `from random import _os as x` は許可モジュール(random)に見えるが、
    # 危険モジュール os を別名 x に束縛してしまい AST 検査をすり抜ける。
    for item in fromlist or ():
        if isinstance(item, str) and item.lstrip("_") in _DANGEROUS_NAMES:
            raise ImportError(
                f"禁止されたモジュール/名前の import は許可されていません: {item}"
            )
    return _b.__import__(name, globals, locals, fromlist, level)


def _make_safe_builtins() -> dict:
    import builtins as _b

    safe = {
        name: getattr(_b, name)
        for name in (_ALLOWED_BUILTINS | _ALLOWED_EXCEPTIONS)
        if hasattr(_b, name)
    }
    safe["__import__"] = _safe_import
    # class 文を許可するために必要（脱獄経路ではない）
    safe["__build_class__"] = _b.__build_class__
    safe["__name__"] = "__agent_sandbox__"
    return safe


def precheck_step_code(code: str) -> None:
    """exec前の軽量ASTチェック。明白な脱獄（ダンダー属性・禁止名）を弾く。

    import 自体はここでは弾かず、実行時の _safe_import がホワイトリスト判定する
    （モデルが import エラーを観測して自己修正できるようにするため）。
    """
    from app.common import JobError

    try:
        tree = ast.parse(code)
    except SyntaxError as exc:
        raise JobError("AGENT_CODE_SYNTAX", f"生成コードの構文エラー: {exc}") from exc

    def _is_forbidden(identifier: str) -> bool:
        # ダンダー(__builtins__ 等)は一括禁止。先頭の "_" を剥がした名前も照合し、
        # `re._sys` / `random._os` のようなプライベート別名経由の脱獄を防ぐ。
        if identifier.startswith("__"):
            return True
        if identifier in _FORBIDDEN_NAMES:
            return True
        return identifier.lstrip("_") in _FORBIDDEN_NAMES

    for node in ast.walk(tree):
        if isinstance(node, ast.Attribute):
            # `.compile` 属性は re.compile 等の安全用途なので許可（ビルトイン
            # compile の Name 参照は引き続き禁止）。
            if node.attr != "compile" and _is_forbidden(node.attr):
                raise JobError(
                    "AGENT_CODE_REJECTED",
                    f"禁止された属性へのアクセスです: {node.attr}",
                )
        if isinstance(node, ast.Name) and _is_forbidden(node.id):
            raise JobError(
                "AGENT_CODE_REJECTED", f"禁止された名前の参照: {node.id}"
            )
        # import の対象名・別名も検査（例: `from datetime import __class__ as x`）
        if isinstance(node, ast.alias):
            if _is_forbidden(node.name) or (
                node.asname and _is_forbidden(node.asname)
            ):
                raise JobError(
                    "AGENT_CODE_REJECTED",
                    f"禁止された名前の import / 別名です: {node.name}",
                )
        # str.format のフィールド経由の属性アクセス脱獄を遮断。
        # 例: '{0._os.environ}'.format(random) / '{0.__class__.__globals__[os]}'。
        # これらは AST では文字列内に隠れて属性チェーンを辿れてしまう
        # （random._os は os 本体、calendar.sys なども露出している）。
        # 置換フィールド `{field!conv:spec}` の field 名部分に `.` があれば属性
        # アクセスと見なして拒否する。`{:.2f}` 等の書式指定は field が空なので素通り。
        if isinstance(node, ast.Constant) and isinstance(node.value, str):
            for field in re.findall(r"\{([^{}]*)\}", node.value):
                head = field.split(":", 1)[0].split("!", 1)[0]
                if "." in head:
                    raise JobError(
                        "AGENT_CODE_REJECTED",
                        "文字列フォーマットのフィールド経由の属性アクセスは禁止されています",
                    )


# worker プロセスに残す環境変数（これ以外は秘密情報の露出を防ぐため削除）。
# Python の起動・ファイル探索・ロケール・一時ディレクトリに必要な OS 標準の集合。
# ※ Windows(本番サーバ)で worker が壊れないよう、Windows 標準変数も保持する。
#   ここに無い独自変数（APIキー等）だけが削除される。
_ENV_KEEP = {
    # 共通 / POSIX
    "PATH", "PYTHONPATH", "PYTHONHOME", "PYTHONIOENCODING", "PYTHONHASHSEED",
    "LANG", "LANGUAGE", "LC_ALL", "LC_CTYPE", "TZ", "HOME", "USER", "LOGNAME",
    "TMPDIR", "TEMP", "TMP",
    # Windows 標準
    "SYSTEMROOT", "SYSTEMDRIVE", "WINDIR", "PATHEXT", "COMSPEC",
    "USERPROFILE", "APPDATA", "LOCALAPPDATA", "HOMEDRIVE", "HOMEPATH",
    "ALLUSERSPROFILE", "PROGRAMDATA", "PROGRAMFILES", "PROGRAMFILES(X86)",
    "PROGRAMW6432", "COMMONPROGRAMFILES", "COMMONPROGRAMFILES(X86)",
    "PUBLIC", "OS", "NUMBER_OF_PROCESSORS", "PROCESSOR_ARCHITECTURE",
    "PROCESSOR_IDENTIFIER", "PROCESSOR_LEVEL", "PROCESSOR_REVISION",
    "USERDOMAIN", "COMPUTERNAME", "SESSIONNAME",
}


def _scrub_env(environ) -> None:
    """秘密情報が紛れ込まないよう、環境変数を最小限に絞る。"""
    for key in list(environ.keys()):
        if key.upper() not in _ENV_KEEP:
            try:
                del environ[key]
            except Exception:
                pass


def _disable_network() -> None:
    """worker からのネットワーク発信を無効化する（移植性あり）。

    クラスを関数で差し替えると `isinstance(x, socket.socket)` 型チェックが壊れて
    openpyxl 等を巻き込む恐れがあるため、`__init__` を潰してインスタンス化のみを
    禁止する（既存接続=mpのpipeは無傷）。名前解決(DNS)も塞ぐ。

    ※ そもそも生成コードは `socket` / `_socket` を import できない
      （import ホワイトリスト外）ため、これは worker プロセス自体に対する
      多層防御の一枚。完全なegress遮断はOSレベル(コンテナ/FW)で行うのが本筋。
    """
    import socket

    def _blocked(*args, **kwargs):
        raise OSError("network access is disabled in the sandbox")

    def _blocked_init(self, *args, **kwargs):
        raise OSError("network access is disabled in the sandbox")

    try:
        socket.socket.__init__ = _blocked_init  # type: ignore[method-assign]
    except Exception:
        socket.socket = _blocked  # type: ignore[assignment]  # フォールバック
    # __init__ を迂回して __new__ 等で生成されても egress できないよう、
    # 接続/送信メソッド自体も潰す（既存の mp pipe は connect/sendto を使わない）。
    for _meth in ("connect", "connect_ex", "sendto"):
        try:
            setattr(socket.socket, _meth, _blocked)
        except Exception:
            pass
    # 高水準API・名前解決を塞ぐ（DNS exfiltration 等の経路も断つ）。
    # getaddrinfo 系だけでなく旧来の resolver(gethostby*) も含める。
    for _name in (
        "create_connection",
        "create_server",
        "getaddrinfo",
        "getnameinfo",
        "gethostbyname",
        "gethostbyname_ex",
        "gethostbyaddr",
        "getfqdn",
    ):
        if hasattr(socket, _name):
            try:
                setattr(socket, _name, _blocked)
            except Exception:
                pass


def _apply_resource_limits() -> None:
    """CPU/ファイルサイズ/FD/プロセス数の上限を設定する（POSIXのみ・best-effort）。"""
    import os

    try:
        import resource
    except Exception:
        return  # Windows 等では resource が無いので no-op

    def _int_env(name: str, default: int) -> int:
        # 不正値（空文字/非数値/0以下）は既定値にフォールバックする
        try:
            value = int(os.getenv(name, str(default)))
        except (TypeError, ValueError):
            return default
        return value if value > 0 else default

    def _limit(res_id, value: int) -> None:
        try:
            soft, hard = resource.getrlimit(res_id)
            new = value if hard == resource.RLIM_INFINITY else min(value, hard)
            resource.setrlimit(res_id, (new, hard))
        except Exception:
            pass

    _limit(resource.RLIMIT_CPU, _int_env("XLSX_AGENT_WORKER_CPU_SEC", 300))
    _limit(resource.RLIMIT_FSIZE, _int_env("XLSX_AGENT_WORKER_FSIZE_MB", 100) * 1024 * 1024)
    _limit(resource.RLIMIT_NOFILE, _int_env("XLSX_AGENT_WORKER_NOFILE", 256))
    if hasattr(resource, "RLIMIT_NPROC"):
        _limit(resource.RLIMIT_NPROC, _int_env("XLSX_AGENT_WORKER_NPROC", 64))
    if os.getenv("XLSX_AGENT_WORKER_MEM_MB") and hasattr(resource, "RLIMIT_AS"):
        mem_mb = _int_env("XLSX_AGENT_WORKER_MEM_MB", 0)  # 既定は無制限（誤検知防止）
        if mem_mb > 0:
            _limit(resource.RLIMIT_AS, mem_mb * 1024 * 1024)


def _harden_worker_process() -> None:
    """worker 子プロセスに OS レベルの隔離を適用する（ベストエフォート）。

    生成コードは別プロセス＋制限付きビルトイン＋AST/import ガードで実行されるが、
    万一それらを抜けられても実害が出ないよう、プロセス自体の
    ネットワーク・環境変数・リソースを絞る。各処理は失敗しても無視し、
    未対応OS（Windows 等）では該当部分が no-op になる。

    ※ import 完了後に呼ぶこと（環境変数スクラブが起動時 import に影響しないよう）。
    """
    import os

    if os.getenv("XLSX_AGENT_DISABLE_NETWORK", "1") != "0":
        try:
            _disable_network()
        except Exception:
            pass
    # リソース上限は XLSX_AGENT_WORKER_* を参照するので、環境変数スクラブより
    # 「先に」実行する（先にスクラブすると設定が消える）。
    try:
        _apply_resource_limits()
    except Exception:
        pass
    # 設定値を読み終えた後に env をスクラブ（XLSX_AGENT_* 自体も worker から消す）
    try:
        _scrub_env(os.environ)
    except Exception:
        pass


def _agent_worker(input_path: str, sheet_name: str, conn) -> None:
    """spawnされる子プロセス本体。wb/wsを保持し、コマンドを逐次処理する。

    ※ `app.main` ではなく `app.common`（副作用なし）から部品を import する。
      `app.main` を子プロセスで読むと `create_app()` が走り JobService の
      スレッドが余計に起動してしまうため。

    各ステップはトランザクション的に扱う：実行が失敗したら、その途中変更を
    破棄して直近の成功状態（committed スナップショット）へロールバックする。
    これにより失敗ステップの中途半端な変更が保存されない。
    """
    try:
        from openpyxl import load_workbook

        from app.common import (
            _close_workbook,
            _keep_vba_for,
            _safe_set_merged_value,
            validate_excel_file,
        )

        # import 完了後に OS レベルの隔離を適用（ネットワーク遮断・env スクラブ・
        # POSIXのリソース上限）。未対応OSでは該当部分が no-op。
        _harden_worker_process()

        keep_vba = _keep_vba_for(input_path)

        def _build_namespace(workbook):
            worksheet = workbook[sheet_name]
            return {
                "__builtins__": _make_safe_builtins(),
                "wb": workbook,
                "ws": worksheet,
                "helpers": {
                    "safe_set_merged_value": lambda merged_range, value: (
                        _safe_set_merged_value(worksheet, merged_range, value)
                    )
                },
            }

        def _snapshot(workbook) -> bytes:
            buffer = io.BytesIO()
            workbook.save(buffer)
            return buffer.getvalue()

        reserved_keys = {"__builtins__", "wb", "ws", "helpers"}
        wb = load_workbook(input_path, keep_vba=keep_vba, data_only=False)
        namespace: dict[str, Any] = _build_namespace(wb)
        committed = _snapshot(wb)  # 直近の「成功状態」のスナップショット
        committed_vars: dict[str, Any] = {}  # 直近成功時のLLM定義変数
        conn.send({"ok": True})
    except Exception:
        try:
            conn.send({"ok": False, "error": traceback.format_exc()[-MAX_OBSERVATION_CHARS:]})
        except Exception:
            pass
        return

    try:
      while True:
        try:
            msg = conn.recv()
        except EOFError:
            break
        cmd = msg.get("cmd")
        if cmd == "exec":
            buf = io.StringIO()
            try:
                with redirect_stdout(buf):
                    exec(compile(msg["code"], "<agent-step>", "exec"), namespace)
                committed = _snapshot(namespace["wb"])  # 成功 → コミット
                # 成功時のLLM定義変数も控えておく（ロールバック時に復元するため）。
                # ただし openpyxl 由来のオブジェクト（Cell/Worksheet/Workbook 等）は
                # ロールバックで破棄される古い wb を指してしまうので保持しない。
                committed_vars = {
                    k: v
                    for k, v in namespace.items()
                    if k not in reserved_keys
                    and not type(v).__module__.startswith("openpyxl")
                }
                conn.send({"ok": True, "stdout": buf.getvalue()[-MAX_OBSERVATION_CHARS:]})
            except Exception:
                error_text = traceback.format_exc()[-MAX_OBSERVATION_CHARS:]
                # 失敗ステップの途中変更を破棄し、直近コミット状態へロールバック。
                # 古い wb は閉じて FD/メモリのリークを防ぐ。
                try:
                    _close_workbook(namespace["wb"])
                except Exception:
                    pass
                try:
                    wb = load_workbook(
                        io.BytesIO(committed), keep_vba=keep_vba, data_only=False
                    )
                    namespace = _build_namespace(wb)
                    # 直近成功時のLLM定義変数を復元（NameError防止）
                    namespace.update(committed_vars)
                except Exception:
                    pass
                conn.send(
                    {
                        "ok": False,
                        "stdout": buf.getvalue()[-MAX_OBSERVATION_CHARS:],
                        "error": error_text,
                    }
                )
        elif cmd == "save":
            target = Path(msg["path"])
            tmp_path = target.with_name(f"{target.stem}.tmp{target.suffix}")
            try:
                namespace["wb"].save(tmp_path)
                validate_excel_file(tmp_path)
                tmp_path.replace(target)
                conn.send({"ok": True})
            except Exception as exc:
                # 検証失敗などで残った一時ファイルを掃除する
                try:
                    if tmp_path.exists():
                        tmp_path.unlink()
                except Exception:
                    pass
                error_code = getattr(exc, "error_code", "EXCEL_SAVE_VALIDATION_FAILED")
                message = getattr(exc, "message", None) or traceback.format_exc()[
                    -MAX_OBSERVATION_CHARS:
                ]
                conn.send({"ok": False, "error_code": error_code, "error": message})
        elif cmd == "close":
            break
    finally:
        # 終了時に wb を確実にクローズして FD/メモリのリークを防ぐ
        try:
            _close_workbook(namespace["wb"])
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


class AgentSandbox:
    """親プロセス側のサンドボックス制御。1つの子プロセスにコマンドを送る。"""

    def __init__(self, input_path: Path, sheet_name: str, init_timeout: int = 30) -> None:
        from app.common import JobError

        ctx = mp.get_context("spawn")
        self._conn, child_conn = ctx.Pipe()
        self._proc = ctx.Process(
            target=_agent_worker,
            args=(str(input_path), sheet_name, child_conn),
            daemon=True,
        )
        self._proc.start()
        child_conn.close()
        if not self._conn.poll(init_timeout):
            self.close()
            raise JobError("AGENT_INIT_FAILED", "サンドボックス初期化がタイムアウトしました")
        ready = self._recv("AGENT_INIT_FAILED", "サンドボックス初期化中にプロセスが予期せず終了しました")
        if not ready.get("ok"):
            self.close()
            raise JobError(
                "AGENT_INIT_FAILED",
                f"サンドボックス初期化に失敗しました: {ready.get('error', '')}",
            )

    def _send(self, message: dict, fail_message: str) -> None:
        from app.common import JobError

        try:
            self._conn.send(message)
        except OSError as exc:
            self.close()
            raise JobError("AGENT_SANDBOX_CRASHED", fail_message, retryable=True) from exc

    def _recv(self, error_code: str, fail_message: str) -> dict:
        from app.common import JobError

        try:
            return self._conn.recv()
        except (EOFError, OSError) as exc:
            self.close()
            raise JobError(error_code, fail_message, retryable=True) from exc

    def run(self, code: str, timeout: int) -> dict:
        from app.common import JobError

        self._send({"cmd": "exec", "code": code}, "サンドボックスへのコマンド送信に失敗しました")
        if not self._conn.poll(timeout):
            self.close()
            raise JobError(
                "EXEC_TIMEOUT", "ステップ実行がタイムアウトしました", retryable=True
            )
        return self._recv(
            "AGENT_SANDBOX_CRASHED",
            "ステップ実行中にサンドボックスプロセスが予期せず終了しました",
        )

    def save(self, path: str, timeout: int) -> dict:
        from app.common import JobError

        self._send({"cmd": "save", "path": path}, "サンドボックスへの保存コマンド送信に失敗しました")
        if not self._conn.poll(timeout):
            self.close()
            raise JobError(
                "EXEC_TIMEOUT", "保存がタイムアウトしました", retryable=True
            )
        return self._recv(
            "AGENT_SANDBOX_CRASHED", "保存中にサンドボックスプロセスが予期せず終了しました"
        )

    def close(self) -> None:
        try:
            if self._proc.is_alive():
                try:
                    self._conn.send({"cmd": "close"})
                except Exception:
                    pass
                self._proc.join(timeout=3)
            if self._proc.is_alive():
                self._proc.terminate()
                self._proc.join(timeout=3)
        except Exception:
            pass
        finally:
            try:
                self._conn.close()
            except Exception:
                pass


# コードフェンスの揺れに対応する。小型モデルは ``` の代わりに ''' や ~~~ を使ったり、
# 言語名を付けなかったり、閉じフェンスを忘れたりする。これらを malformed 扱いにすると
# コードを書いているのに延々と手を浪費するため、フェンス種別と閉じ忘れを許容する。
_FENCE = r"(?:```|'''|~~~)"


def _is_vacuous_reply(text: str) -> bool:
    """応答が実体を持たない（フェンス記号や空白だけ・空）かを判定する。

    小型モデルは編集後 DONE を出さず、閉じフェンス ``` だけ／空応答を返して空転する。
    これは「実質完了」とみなしてよいが、普通の散文（続行の意図があるかもしれない）は
    対象外にする。判定はフェンス文字と空白を取り除いて何も残らない場合のみ True。
    """
    return re.sub(r"[`'~\s]", "", text or "") == ""


def _strip_done_lines(block: str) -> tuple[str, bool]:
    """ブロックから単独行の DONE を取り除き、(残り, DONEがあったか) を返す。
    DONE 判定は単独行のみ（"not done" 等の部分一致を誤判定しないため）。"""
    kept: list[str] = []
    found = False
    for line in block.splitlines():
        if line.strip().upper().rstrip(".!。 ") == "DONE":
            found = True
            continue
        kept.append(line)
    return "\n".join(kept).strip(), found


def parse_action(text: str) -> tuple[str, str]:
    """LLM応答を (種別, コード) に解釈する。
    種別は 'code' / 'code_done' / 'done' / 'malformed'。

    'code_done' は「コード＋（外側に）単独行 DONE」の応答で、1応答で実行と完了を兼ねる。
    ``` だけでなく ''' / ~~~、閉じフェンス忘れ、フェンス無しのベタ書きにも対応する。
    """
    text = text or ""
    # 1) 閉じフェンスありのコードブロック（``` ''' ~~~ いずれも・言語名は任意）
    match = re.search(
        rf"{_FENCE}[ \t]*(?:python|py)?[ \t]*\r?\n(.*?){_FENCE}",
        text,
        flags=re.DOTALL | re.IGNORECASE,
    )
    if not match:
        # 2) 閉じフェンスを付け忘れた場合は、開きフェンス以降を全部コードとみなす
        match = re.search(
            rf"{_FENCE}[ \t]*(?:python|py)?[ \t]*\r?\n(.*)\Z",
            text,
            flags=re.DOTALL | re.IGNORECASE,
        )
    code = ""
    if match and match.group(1).strip():
        code = match.group(1).strip()
        # "```\npython\n..." のように言語名が独立行で来るケースを除去
        code = re.sub(r"\A(?:python|py)[ \t]*\r?\n", "", code, flags=re.IGNORECASE)

    # コードブロック外のテキストで DONE を判定（閉じ忘れ分も含めフェンス区間を除去）
    outside = re.sub(rf"{_FENCE}.*?{_FENCE}", " ", text, flags=re.DOTALL)
    outside = re.sub(rf"{_FENCE}.*\Z", " ", outside, flags=re.DOTALL)
    has_done = any(
        line.strip().upper().rstrip(".!。 ") == "DONE" for line in outside.splitlines()
    )
    if code:
        # 閉じ忘れ抽出だと DONE がコード末尾に混ざることがあるので分離する
        code, done_in_code = _strip_done_lines(code)
        if code:
            return ("code_done", code) if (has_done or done_in_code) else ("code", code)
    # 3) フェンスが全く無くても、openpyxl 操作らしき記述があれば実行を試みる
    #    （散文・思考は除外しつつ、ベタ書きコードを malformed で捨てない）。
    #    "コード\nDONE" のようにベタ書きコードと DONE が同居しても取りこぼさないよう、
    #    この判定は下の `has_done` 早期returnより前に行う。
    if re.search(
        r"(?m)^\s*(?:from\s+openpyxl|import\s+\w|ws\s*[\[.=]|wb\s*[\[.=]|"
        r"for\s+\w+\s+in\b|helpers\b)",
        text,
    ):
        body, done_in_code = _strip_done_lines(text.strip())
        if body:
            return ("code_done", body) if (has_done or done_in_code) else ("code", body)
    if has_done:
        return "done", ""
    # コードも明示的な DONE も無い → 不正応答。完了扱いにせず再試行させる。
    return "malformed", ""


def run_agent(
    working_path: Path,
    sheet_name: str,
    instruction: str,
    summary: dict[str, Any],
    llm,
    max_steps: int = 6,
    step_timeout: int = 30,
    think: bool | None = None,
    progress=None,
    log_path: Path | None = None,
) -> dict:
    """ReActループ本体。working_path を直接編集する。失敗時 JobError。

    progress(label:str) があれば各段階で呼ばれ、UIに進捗を出せる。
    log_path を渡すと、各手の生のLLM応答・実行結果をそのファイルへ追記する
    （何が起きて手が伸びたかを後から追える。タイムアウトでも途中まで残るよう即flush）。
    """

    def _say(label: str) -> None:
        if progress:
            try:
                progress(label)
            except Exception:
                pass

    def _log(msg: str) -> None:
        if not log_path:
            return
        try:
            with open(log_path, "a", encoding="utf-8") as fh:
                fh.write(msg + "\n")
                fh.flush()
        except Exception:
            pass

    from datetime import datetime

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _log("=" * 70)
    _log(f"[{ts}] 指示: {instruction!r}")
    _log(
        f"model={getattr(llm, 'model', '?')} think={think} "
        f"max_steps={max_steps} sheet={sheet_name!r}"
    )
    _log("-" * 70)

    from app.common import JobError

    sandbox = AgentSandbox(working_path, sheet_name)
    transcript: list[dict[str, Any]] = []
    applied = 0
    completed = False
    try:
        for step in range(1, max_steps + 1):
            _say(f"AIが編集コードを生成中…(手{step})")
            text = llm.agent_step(summary, instruction, transcript, think=think)
            kind, code = parse_action(text)
            _log(f"--- 手{step} (kind={kind}) ---")
            # ログはローカルのデバッグ用途なので丸めない（LLMへ渡す transcript のみ
            # MAX_OBSERVATION_CHARS で制限する）。thinking出力や長文でも全文残す。
            _log("[LLM応答]\n" + (text or "").strip())
            if kind == "done":
                completed = True
                _log("[結果] DONE（完了宣言）")
                break
            if kind == "malformed":
                # すでに有効な編集が1つ以上適用済みで、追加応答が「実体なし」
                # （閉じフェンス ``` だけ・空応答 等）の場合に限り、DONE を出さない
                # 小型モデルが「実質完了」している状況とみなして保存する。ここでループを
                # 続けても手を浪費し、最終的にステップ上限で編集ごと破棄されてしまう
                # ため、得られた成果を残す方が良い（続きはチャットの次の指示で足せる）。
                # 普通の散文（続行/リカバリの意図があり得る）は早期完了させず再試行する。
                if applied > 0 and _is_vacuous_reply(text):
                    completed = True
                    _log(
                        f"[結果] 実体なし応答だが適用済み編集あり → 完了として保存 "
                        f"(applied={applied})"
                    )
                    break
                # まだ編集が無い、または散文での不正応答 → コード or DONE を促して再試行
                _log("[結果] MALFORMED（コードもDONEも無い不正応答）")
                transcript.append(
                    {
                        "step": step,
                        "code": "",
                        "observation": "MALFORMED: ```python のコードブロック1つ、"
                        "または完了なら DONE のみを返してください。",
                    }
                )
                continue
            _say("生成コードを確認中…")
            try:
                precheck_step_code(code)
            except JobError as rejected:
                _log(f"[結果] REJECTED（安全チェック）: {rejected.message}")
                transcript.append(
                    {
                        "step": step,
                        "code": code,
                        "observation": f"REJECTED: {rejected.message}",
                    }
                )
                continue
            _say("編集を実行中…")
            result = sandbox.run(code, timeout=step_timeout)
            if result.get("ok"):
                applied += 1
                stdout = (result.get("stdout") or "").strip()
                observation = "OK" + (f"\n{stdout}" if stdout else "")
            else:
                observation = "ERROR:\n" + (
                    result.get("error") or result.get("stdout") or "unknown error"
                )
            _log("[実行結果] " + observation)  # ログは全文（長いトレースも切らない）
            transcript.append(
                {
                    "step": step,
                    "code": code,
                    "observation": observation[:MAX_OBSERVATION_CHARS],
                }
            )
            # コード＋DONE を1回で返した応答は、成功したらそのまま完了にする
            # （完了確認の追加LLM呼び出しを省いて高速化）。失敗時は通常どおり次手で修正。
            if kind == "code_done" and result.get("ok"):
                completed = True
                _log("[結果] コード＋DONEで実行成功 → 完了")
                break

        # 1件も編集できていない場合だけ、再試行可能なエラーにする。
        #   ・completed（DONE宣言）なのに applied==0 → 中身のない完了 → NO_CHANGES
        #   ・未完了かつ applied==0 → ステップ上限・何も成果なし → STEP_LIMIT
        if applied == 0:
            if completed:
                _log("[終了] 失敗: 有効な編集なし（AGENT_NO_CHANGES）")
                raise JobError(
                    "AGENT_NO_CHANGES",
                    "エージェントが有効な編集を生成できませんでした",
                    retryable=True,
                )
            _log(f"[終了] 失敗: ステップ上限({max_steps})到達・編集なし")
            raise JobError(
                "AGENT_STEP_LIMIT",
                f"ステップ上限({max_steps})に達しました。指示を分割するか "
                "XLSX_AGENT_MAX_STEPS を増やして再実行してください。",
                retryable=True,
            )
        # ここに来るのは「編集が1件以上ある」場合。たとえ DONE 未宣言（未完了）でも、
        # 弱いモデルは完了を明示できないだけのことが多く、得られた成果を破棄するより
        # 保存する方がユーザーに有益。未完了の可能性は completed フラグで呼び出し側へ伝える。
        if not completed:
            _log(f"[終了] 未完了だが適用済み編集を保存 (applied={applied})")
        save_result = sandbox.save(str(working_path), timeout=step_timeout)
        if not save_result.get("ok"):
            _log(f"[終了] 失敗: 保存エラー {save_result.get('error', '')}")
            raise JobError(
                save_result.get("error_code", "EXCEL_SAVE_VALIDATION_FAILED"),
                f"保存に失敗しました: {save_result.get('error', '')}",
            )
        _log(
            f"[終了] {'成功' if completed else '部分保存(未完了)'}: "
            f"steps={len(transcript)} applied={applied}"
        )
    finally:
        sandbox.close()
    return {
        "steps": len(transcript),
        "applied": applied,
        "transcript": transcript,
        "completed": completed,
    }
