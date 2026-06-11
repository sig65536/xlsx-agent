"""副作用のない共通ユーティリティ。

`app.main` を import すると末尾の `app = create_app()` が走り JobService（＝
バックグラウンドスレッド）が起動してしまう。サンドボックス子プロセスなどから
安全に使えるよう、FastAPI/JobService に依存しない部品だけをここに切り出す。
"""

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class JobError(Exception):
    def __init__(self, error_code: str, message: str, retryable: bool = False):
        self.error_code = error_code
        self.message = message
        self.retryable = retryable
        super().__init__(message)


def _close_workbook(wb) -> None:
    vba_archive = getattr(wb, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()
        wb.vba_archive = None
    wb.close()


def _keep_vba_for(path) -> bool:
    """VBA保持の要否を拡張子から決める。

    .xlsm のみ keep_vba=True にする。plain .xlsx を keep_vba=True で保存すると
    openpyxl がブックの content-type を macroEnabled (=.xlsm用) にしてしまい、
    拡張子 .xlsx と矛盾して Excel が「破損」と判断する。拡張子に合わせるのが正解。
    """
    return str(path).lower().endswith(".xlsm")


def _workbook_content_type(path: Path) -> str:
    """[Content_Types].xml から /xl/workbook.xml のContentTypeを取り出す。

    属性の順序や空白に依存しないよう、正規表現ではなくXMLとしてパースする。
    """
    with zipfile.ZipFile(path) as zf:
        content_types_xml = zf.read("[Content_Types].xml")
    try:
        root = ET.fromstring(content_types_xml)
    except ET.ParseError:
        return ""
    for override in root.iter():
        if override.tag.endswith("}Override") or override.tag == "Override":
            if override.get("PartName") == "/xl/workbook.xml":
                return override.get("ContentType", "")
    return ""


def validate_excel_file(path: Path) -> None:
    """保存後のxlsx/xlsmが壊れていないか検証する。

    壊れたファイルをそのまま結果として渡さないための最終チェック。
    - zip構造（必須パートの存在）
    - 拡張子と content-type の一致（.xlsx なのに macroEnabled になっていない等）
    - openpyxl での再読込

    ※ content-type の不一致は openpyxl では開けてしまうため、Excel だけが弾く
      「開けないファイル」を検出するにはこのチェックが必須。
    """
    ext = path.suffix.lower()
    keep_vba = ext == ".xlsm"
    expect_macro = ext == ".xlsm"

    if not path.exists():
        raise JobError("EXCEL_SAVE_VALIDATION_FAILED", "出力ファイルが存在しません")
    if path.stat().st_size == 0:
        raise JobError("EXCEL_SAVE_VALIDATION_FAILED", "出力ファイルが0バイトです")
    if not zipfile.is_zipfile(path):
        raise JobError(
            "EXCEL_SAVE_VALIDATION_FAILED", "出力ファイルがzip形式ではありません"
        )
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
        for required in ("[Content_Types].xml", "xl/workbook.xml"):
            if required not in names:
                raise JobError(
                    "EXCEL_SAVE_VALIDATION_FAILED", f"{required} がありません"
                )
        workbook_ct = _workbook_content_type(path)
        is_macro_ct = "macroEnabled" in workbook_ct
        if expect_macro and not is_macro_ct:
            raise JobError(
                "EXCEL_SAVE_VALIDATION_FAILED",
                ".xlsm なのにマクロ有効ブックのcontent-typeになっていません",
            )
        if not expect_macro and is_macro_ct:
            raise JobError(
                "EXCEL_SAVE_VALIDATION_FAILED",
                ".xlsx なのにマクロ有効ブックのcontent-typeになっています"
                "（keep_vba誤用によるExcel破損）",
            )
    except JobError:
        raise
    except Exception as exc:
        raise JobError(
            "EXCEL_SAVE_VALIDATION_FAILED", f"zip検証に失敗しました: {exc}"
        ) from exc
    try:
        wb = load_workbook(path, keep_vba=keep_vba, data_only=False)
        _close_workbook(wb)
    except Exception as exc:
        raise JobError(
            "EXCEL_SAVE_VALIDATION_FAILED", f"openpyxl再読込に失敗しました: {exc}"
        ) from exc


def _safe_set_merged_value(ws, merged_range: str, value: Any) -> None:
    merged_range = merged_range.upper()  # LLMが小文字で渡しても動くよう正規化
    ws.unmerge_cells(merged_range)
    anchor = merged_range.split(":", 1)[0]
    ws[anchor] = value
    ws.merge_cells(merged_range)
