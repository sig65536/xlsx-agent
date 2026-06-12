import ast
import io
import json
import multiprocessing as mp
import os
import queue
import re
import secrets
import shutil
import threading
import time
import traceback
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import ProxyHandler, Request, build_opener

from fastapi import FastAPI, File, Form, HTTPException, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from starlette.concurrency import run_in_threadpool

from app.common import (
    JobError,
    _close_workbook,
    _keep_vba_for,
    _safe_set_merged_value,
    validate_excel_file,
)

STATIC_DIR = Path(__file__).resolve().parent / "static"
CONFIG_ENV_PATH = Path(__file__).resolve().parent.parent / "config.env"


def _load_config_env(path: Path | None = None, env: dict | None = None) -> None:
    """config.env を読み、未設定の環境変数だけ補う。

    モデル名などの設定を「config.env 一箇所」で切り替えられるようにするための仕組み。
    優先順位は 実環境変数 > config.env > コード内の既定値。
    """
    path = path or CONFIG_ENV_PATH
    env = env if env is not None else os.environ
    if not path.exists():
        return
    try:
        for raw in path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and key not in env:
                env[key] = val
    except Exception:
        pass


_load_config_env()

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024
JOB_TTL_SECONDS = 3600
LLM_MAX_RETRY = 3
# 編集方式: "agent"（ReActループ・既定）/ "oneshot"（単発コード生成）
AGENT_DEFAULT_MODE = os.getenv("XLSX_AGENT_MODE", "agent").lower()
AGENT_MAX_STEPS = int(os.getenv("XLSX_AGENT_MAX_STEPS", "6"))
AGENT_STEP_TIMEOUT = int(os.getenv("XLSX_AGENT_STEP_TIMEOUT", "30"))
SUMMARY_FORMULA_MAX_ROWS = 200
SUMMARY_FORMULA_MAX_COLS = 50
PREVIEW_MAX_CHANGED_CELLS = 500
PREVIEW_FORMULA_NOTE = "数式セルは文字列比較です。実値はExcelで再計算されます。"


class JobStatus:
    QUEUED = "queued"
    ANALYZING = "analyzing"
    GENERATING = "generating"
    CHECKING = "checking"
    EXECUTING = "executing"
    PREVIEW_READY = "preview_ready"
    APPROVED = "approved"
    DONE = "done"
    ERROR = "error"


@dataclass
class Job:
    job_id: str
    instruction: str
    sheet_name: str | None
    created_at: float
    work_dir: Path
    source_path: Path
    status: str = JobStatus.QUEUED
    error_code: str | None = None
    message: str | None = None
    retryable: bool = False
    preview: dict[str, Any] | None = None
    result_path: Path | None = None
    download_token: str | None = None


# Ollama はローカル(127.0.0.1)なので、システム/環境のプロキシを通さない opener を使う。
# 社内プロキシやセキュリティソフトが localhost を横取りして 403/HTML を返すのを防ぐ
# （Ollama CLI はループバックをプロキシ除外するが、urllib は既定でプロキシ経由になる）。
_NO_PROXY_OPENER = build_opener(ProxyHandler({}))


class LLMClient:
    def __init__(self) -> None:
        self.endpoint = os.getenv(
            # localhost は Windows で IPv6(::1) に解決され Ollama(IPv4)へ繋がらず
            # 10061(接続拒否)になることがあるため、既定は 127.0.0.1 を使う。
            "OLLAMA_ENDPOINT", "http://127.0.0.1:11434/api/generate"
        )
        self.model = os.getenv("OLLAMA_MODEL", "gemma4-e4b:latest")
        self.timeout = int(os.getenv("LLM_TIMEOUT_SECONDS", "60"))
        # thinking(推論)モード。gemma4 等の対応モデルで思考トレースを有効化する。
        # CPU推論では遅くなるため既定はオフ。Ollamaは thinking を response と分離する。
        self.think = os.getenv("XLSX_AGENT_THINK", "0").lower() not in ("0", "false", "")
        self._resolved_model: str | None = None

    @property
    def _tags_endpoint(self) -> str:
        base = self.endpoint.rsplit("/api/", 1)[0]
        return f"{base}/api/tags"

    def _list_models(self, timeout: int | None = None) -> list[str]:
        req = Request(self._tags_endpoint, method="GET")
        with _NO_PROXY_OPENER.open(req, timeout=timeout or self.timeout) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        return [m.get("name", "") for m in payload.get("models", []) if m.get("name")]

    def diagnose(self) -> dict[str, Any]:
        """Ollama 疎通とモデルの存在を素早く確認する（/healthz 用）。"""
        info: dict[str, Any] = {"model": self.model, "think": self.think}
        try:
            installed = self._list_models(timeout=5)
            info["ollama"] = "ok"
            info["installed_models"] = installed
            info["resolved_model"] = self.resolve_model()
            info["model_found"] = info["resolved_model"] in installed
        except Exception as exc:
            info["ollama"] = f"unreachable: {exc}"
        return info

    def resolve_model(self) -> str:
        """設定モデル名を、実際にOllamaへ入っているタグ名に解決する。

        `ollama pull gemma4` は `gemma4:latest`、`ollama pull gemma4:e4b` は
        `gemma4:e4b` という名前で保存されるため、設定値とインストール名が
        食い違うと404になる。ベース名(gemma4)が一致するタグへ自動で寄せる。
        """
        if self._resolved_model:
            return self._resolved_model
        try:
            installed = self._list_models()
        except (URLError, OSError, ValueError):
            return self.model  # 取得失敗時は設定値のまま試す
        if not installed or self.model in installed:
            self._resolved_model = self.model
            return self.model
        base = self.model.split(":", 1)[0]
        for candidate in (f"{base}:e4b", f"{base}:latest", base):
            if candidate in installed:
                self._resolved_model = candidate
                return candidate
        same_base = [m for m in installed if m.split(":", 1)[0] == base]
        self._resolved_model = same_base[0] if same_base else self.model
        return self._resolved_model

    def _call_ollama(self, prompt: str, think: bool | None = None) -> str:
        """Ollama の generate API を叩き、response テキストを返す。"""
        model = self.resolve_model()
        body = {
            "model": model,
            "stream": False,
            "prompt": prompt,
            "options": {"temperature": 0.1},
        }
        if self.think if think is None else think:
            body["think"] = True  # 思考トレースは payload["thinking"] に分離される
        data = json.dumps(body).encode("utf-8")
        req = Request(
            self.endpoint,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with _NO_PROXY_OPENER.open(req, timeout=self.timeout) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
        except HTTPError as exc:
            detail = ""
            try:
                detail = exc.read().decode("utf-8")[:300]
            except Exception:  # pragma: no cover - 補助的な詳細取得のみ
                pass
            if exc.code == 404:
                raise JobError(
                    "LLM_MODEL_NOT_FOUND",
                    f"Ollamaにモデル '{model}' が見つかりません。"
                    f"サーバーで `ollama pull {self.model}` を実行してください。{detail}",
                ) from exc
            if exc.code == 403:
                raise JobError(
                    "LLM_FORBIDDEN",
                    "Ollamaが403を返しました。多くは (1) localhost通信を横取りする"
                    "プロキシ/セキュリティソフト、または (2) OllamaのCORS保護が原因です。"
                    "(1)は本アプリがプロキシを除外して呼ぶので解消されるはず。残る場合は"
                    "Ollama側で OLLAMA_ORIGINS=* を設定して再起動してください"
                    "（OLLAMA_HOST は変更しない）。"
                    f" {detail}",
                ) from exc
            raise JobError(
                "LLM_HTTP_ERROR",
                f"LLM呼び出しがHTTP {exc.code} を返しました: {detail}",
                retryable=True,
            ) from exc
        except URLError as exc:
            raise JobError(
                "LLM_TIMEOUT", f"LLM呼び出しに失敗しました: {exc}", retryable=True
            ) from exc
        return payload.get("response", "")

    @staticmethod
    def _extract_code(text: str) -> str:
        match = re.search(r"```(?:python)?\s*(.*?)```", text, flags=re.DOTALL)
        if match:
            return match.group(1).strip()
        return text.strip()

    def generate_code(
        self, summary: dict[str, Any], instruction: str, feedback: str = ""
    ) -> str:
        prompt = (
            "あなたはExcel編集用Pythonコード生成器です。"
            "出力は```python ... ```のコードブロック1つだけにしてください。"
            "説明文、思考過程、Markdown本文、箇条書きは禁止です。"
            "使用可能な変数は wb, ws, helpers のみです。"
            "import / open / exec / eval / ファイルI/O / ネットワーク / OS操作は禁止です。\n"
            f"sheet_summary={json.dumps(summary, ensure_ascii=False)}\n"
            f"instruction={instruction}\n"
            f"feedback={feedback}\n"
        )
        return self._extract_code(self._call_ollama(prompt))

    def agent_step(
        self,
        summary: dict[str, Any],
        instruction: str,
        transcript: list[dict[str, Any]],
        think: bool | None = None,
    ) -> str:
        """ReActエージェントの1ターン分の生応答（コードブロック or DONE）を返す。"""
        history = ""
        for item in transcript[-4:]:
            history += (
                f"\n# step {item['step']} のコード:\n{item['code']}\n"
                f"# 実行結果:\n{item['observation']}\n"
            )
        prompt = (
            "あなたはopenpyxlでExcelを編集するエージェントです。"
            f"変数 wb(Workbook), ws(対象シート '{summary.get('sheet_name', '')}'), "
            "helpers が使えます。\n"
            "各ターンの応答は次のどちらか一つだけ:\n"
            "  1) 次に実行するコードを ```python ... ``` で1ブロック"
            "（print()で値を確認してよい）\n"
            "  2) 指示が完了したら DONE の一語のみ\n"
            "import可能: datetime, re, math, openpyxl.styles 等"
            "（os / sys / ファイル / ネットワークは不可）。"
            "説明・思考は書かない。\n"
            f"指示: {instruction}\n"
            f"シート要約: {json.dumps(summary, ensure_ascii=False)}\n"
            f"これまでの経過:{history if history else ' （なし。最初のステップ）'}\n"
        )
        return self._call_ollama(prompt, think=think)


class CodeChecker:
    FORBIDDEN_NAMES = {
        "__import__",
        "eval",
        "exec",
        "compile",
        "open",
        "os",
        "sys",
        "subprocess",
        "socket",
        "shutil",
        "input",
        "breakpoint",
        "globals",
        "locals",
        "vars",
        "dir",
        "getattr",
        "setattr",
        "delattr",
        "type",
        "object",
        "super",
        "memoryview",
    }
    FORBIDDEN_NODES = (
        ast.FunctionDef,
        ast.AsyncFunctionDef,
        ast.ClassDef,
        ast.Lambda,
        ast.With,
        ast.AsyncWith,
        ast.Try,
        ast.Global,
        ast.Nonlocal,
        ast.Delete,
        ast.Raise,
        ast.Yield,
        ast.YieldFrom,
        ast.Await,
    )

    def validate(self, code: str, non_anchor_cells: set[str]) -> None:
        try:
            tree = ast.parse(code)
        except SyntaxError as exc:
            raise JobError(
                "CODE_CHECK_FAILED", f"生成コードの構文エラー: {exc}"
            ) from exc

        for node in ast.walk(tree):
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                raise JobError(
                    "CODE_CHECK_FAILED",
                    "生成コードに禁止された操作（import）が含まれていました",
                )
            if isinstance(node, self.FORBIDDEN_NODES):
                raise JobError(
                    "CODE_CHECK_FAILED",
                    f"生成コードに禁止された構文が含まれていました: {type(node).__name__}",
                )
            if isinstance(node, ast.Name) and node.id in self.FORBIDDEN_NAMES:
                raise JobError(
                    "CODE_CHECK_FAILED",
                    f"禁止された名前参照が含まれていました: {node.id}",
                )
            if isinstance(node, ast.Attribute) and node.attr.startswith("__"):
                raise JobError(
                    "CODE_CHECK_FAILED", "ダンダー属性アクセスは禁止されています"
                )
            if isinstance(node, ast.Call):
                if (
                    isinstance(node.func, ast.Name)
                    and node.func.id in self.FORBIDDEN_NAMES
                ):
                    raise JobError(
                        "CODE_CHECK_FAILED",
                        f"禁止された関数呼び出しが含まれていました: {node.func.id}",
                    )
            if isinstance(node, (ast.Assign, ast.AugAssign)):
                targets = (
                    node.targets if isinstance(node, ast.Assign) else [node.target]
                )
                for target in targets:
                    addr = self._extract_ws_address(target)
                    if addr and addr in non_anchor_cells:
                        raise JobError(
                            "MERGED_CELL_CONFLICT",
                            f"結合セル非アンカーへの書き込みは禁止です: {addr}",
                        )

    def _extract_ws_address(self, node: ast.AST) -> str | None:
        if (
            isinstance(node, ast.Subscript)
            and isinstance(node.value, ast.Name)
            and node.value.id == "ws"
        ):
            if isinstance(node.slice, ast.Constant) and isinstance(
                node.slice.value, str
            ):
                return node.slice.value.upper()
        if (
            isinstance(node, ast.Call)
            and isinstance(node.func, ast.Attribute)
            and node.func.attr == "cell"
        ):
            if isinstance(node.func.value, ast.Name) and node.func.value.id == "ws":
                row = None
                column = None
                if len(node.args) >= 2:
                    row = self._const_int(node.args[0])
                    column = self._const_int(node.args[1])
                for kw in node.keywords:
                    if kw.arg == "row":
                        row = self._const_int(kw.value)
                    if kw.arg in {"column", "col"}:
                        column = self._const_int(kw.value)
                if row and column:
                    return f"{self._col_letters(column)}{row}"
        return None

    @staticmethod
    def _const_int(node: ast.AST) -> int | None:
        if isinstance(node, ast.Constant) and isinstance(node.value, int):
            return node.value
        return None

    @staticmethod
    def _col_letters(column: int) -> str:
        out = ""
        while column:
            column, rem = divmod(column - 1, 26)
            out = chr(65 + rem) + out
        return out


def _summary_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _row_values(ws, row: int, min_col: int, max_col: int) -> list[Any]:
    return [
        _summary_value(ws.cell(row=row, column=col).value)
        for col in range(min_col, max_col + 1)
    ]


def _non_empty_count(values: list[Any]) -> int:
    return sum(value not in (None, "") for value in values)


def _summarize_sheet(ws) -> dict[str, Any]:
    sample_max_row = min(max(ws.max_row, 1), 30)
    sample_max_col = 20
    sample_values = [
        {"row": row, "values": _row_values(ws, row, 1, sample_max_col)}
        for row in range(1, sample_max_row + 1)
    ]

    merged_cells = []
    for rng in ws.merged_cells.ranges:
        min_col, min_row, _max_col, _max_row = rng.bounds
        anchor = f"{CodeChecker._col_letters(min_col)}{min_row}"
        merged_cells.append(
            {
                "range": str(rng),
                "anchor": anchor,
                "value": _summary_value(ws[anchor].value),
            }
        )

    hidden_rows = [
        idx for idx, dimension in ws.row_dimensions.items() if dimension.hidden
    ]
    hidden_columns = [
        col for col, dimension in ws.column_dimensions.items() if dimension.hidden
    ]

    formulas: list[str] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, SUMMARY_FORMULA_MAX_ROWS),
        min_col=1,
        max_col=min(ws.max_column, SUMMARY_FORMULA_MAX_COLS),
    ):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.coordinate)

    header_candidates = []
    for sample_row in sample_values:
        values = sample_row["values"]
        string_count = sum(
            1 for value in values if isinstance(value, str) and value.strip()
        )
        if string_count >= 2 and _non_empty_count(values) >= 2:
            header_candidates.append({"row": sample_row["row"], "values": values})

    table_like_ranges = []
    for candidate in header_candidates[:5]:
        row = candidate["row"]
        values = candidate["values"]
        non_empty_cols = [
            idx + 1 for idx, value in enumerate(values) if value not in (None, "")
        ]
        if not non_empty_cols:
            continue
        min_col = min(non_empty_cols)
        max_col = max(non_empty_cols)
        end_row = row
        for scan_row in range(row + 1, min(ws.max_row, row + 200) + 1):
            row_values = _row_values(ws, scan_row, min_col, max_col)
            if _non_empty_count(row_values) == 0:
                break
            end_row = scan_row
        table_like_ranges.append(
            {
                "range": f"{CodeChecker._col_letters(min_col)}{row}:{CodeChecker._col_letters(max_col)}{end_row}",
                "header_row": row,
                "columns": values[min_col - 1 : max_col],
            }
        )

    return {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
        "merged_cells": merged_cells,
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_columns,
        "protected": bool(ws.protection.sheet),
        "sample_values": sample_values,
        "header_candidates": header_candidates,
        "table_like_ranges": table_like_ranges,
        "formula_cells": formulas,
    }


def _non_anchor_cells(ws) -> set[str]:
    blocked: set[str] = set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        anchor = f"{CodeChecker._col_letters(min_col)}{min_row}"
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                address = f"{CodeChecker._col_letters(col)}{row}"
                if address != anchor:
                    blocked.add(address)
    return blocked


def _sandbox_runner(path: str, sheet_name: str, code: str, result_queue: Any) -> None:
    try:
        wb = load_workbook(path, keep_vba=_keep_vba_for(path), data_only=False)
        ws = wb[sheet_name]
        safe_globals = {
            "__builtins__": {
                "range": range,
                "len": len,
                "min": min,
                "max": max,
                "sum": sum,
                "enumerate": enumerate,
                "int": int,
                "float": float,
                "str": str,
                "bool": bool,
                "list": list,
                "dict": dict,
                "set": set,
                "tuple": tuple,
                "abs": abs,
                "any": any,
                "all": all,
                "zip": zip,
            }
        }
        safe_locals = {
            "wb": wb,
            "ws": ws,
            "helpers": {
                "safe_set_merged_value": lambda merged_range, value: (
                    _safe_set_merged_value(ws, merged_range, value)
                )
            },
        }
        exec(compile(code, "<generated>", "exec"), safe_globals, safe_locals)
        target = Path(path)
        # 一時ファイルは有効な拡張子を保つ（.tmp だと openpyxl が再読込を拒否する）。
        tmp_path = target.with_name(f"{target.stem}.tmp{target.suffix}")
        wb.save(tmp_path)
        _close_workbook(wb)
        # 壊れたファイルを採用しないよう、整合性検証後に原子的に差し替える。
        validate_excel_file(tmp_path)
        tmp_path.replace(target)
        result_queue.put({"ok": True})
    except JobError as je:
        result_queue.put(
            {"ok": False, "error": je.message, "error_code": je.error_code}
        )
    except Exception:
        result_queue.put({"ok": False, "error": traceback.format_exc()})


def _exec_in_sandbox(
    path: Path, sheet_name: str, code: str, timeout_sec: int = 30
) -> None:
    ctx = mp.get_context("spawn")
    q = ctx.Queue()
    proc = ctx.Process(target=_sandbox_runner, args=(str(path), sheet_name, code, q))
    proc.start()
    proc.join(timeout=timeout_sec)
    if proc.is_alive():
        proc.terminate()
        proc.join(5)
        raise JobError(
            "EXEC_TIMEOUT", "生成コードの実行がタイムアウトしました", retryable=True
        )
    result = q.get() if not q.empty() else {"ok": False, "error": "no result"}
    if not result.get("ok"):
        error_code = result.get("error_code", "EXEC_RUNTIME_ERROR")
        raise JobError(
            error_code,
            f"実行エラー: {result.get('error', '')}",
            retryable=error_code != "EXCEL_SAVE_VALIDATION_FAILED",
        )


def _diff_sheet(ws_before, ws_after, sheet_name: str) -> list[dict[str, Any]]:
    max_row = max(ws_before.max_row, ws_after.max_row)
    max_col = max(ws_before.max_column, ws_after.max_column)
    changes = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            c1 = ws_before.cell(row=row, column=col)
            c2 = ws_after.cell(row=row, column=col)
            if c1.value != c2.value or c1.style_id != c2.style_id:
                coord = f"{CodeChecker._col_letters(col)}{row}"
                changes.append(
                    {
                        "sheet": sheet_name,
                        "cell": coord,
                        "before": c1.value,
                        "after": c2.value,
                        "before_formula": c1.value
                        if isinstance(c1.value, str) and c1.value.startswith("=")
                        else None,
                        "after_formula": c2.value
                        if isinstance(c2.value, str) and c2.value.startswith("=")
                        else None,
                        "style_changed": c1.style_id != c2.style_id,
                    }
                )
    return changes


def _create_preview(
    before_path: Path, after_path: Path, sheet_name: str
) -> dict[str, Any]:
    before = load_workbook(
        before_path, keep_vba=_keep_vba_for(before_path), data_only=False
    )
    after = load_workbook(after_path, keep_vba=_keep_vba_for(after_path), data_only=False)
    changed_cells: list[dict[str, Any]] = []
    notes = [PREVIEW_FORMULA_NOTE]
    try:
        # 全シートを比較する。エージェントは他シートも編集できるため、対象シート
        # だけ差分を取ると「見ていない変更」をユーザーが承認してしまう。
        before_names = set(before.sheetnames)
        after_names = set(after.sheetnames)

        # 対象シートを先頭に、それ以外を後ろに並べて差分を取る
        common = [n for n in after.sheetnames if n in before_names]
        common.sort(key=lambda n: (n != sheet_name, n))
        for name in common:
            changed_cells.extend(_diff_sheet(before[name], after[name], name))

        added = [n for n in after.sheetnames if n not in before_names]
        removed = [n for n in before.sheetnames if n not in after_names]
        if added:
            empty_ws = Workbook().active  # 空シート基準に新規シートの内容を差分表示
            for name in added:
                changed_cells.extend(_diff_sheet(empty_ws, after[name], name))
                notes.append(f"シート「{name}」が新規追加されました。")
        for name in removed:
            notes.append(f"シート「{name}」が削除されました。")
    finally:
        # 差分中に例外が出ても FD を確実に解放する
        _close_workbook(before)
        _close_workbook(after)
    return {
        "sheet_name": sheet_name,
        "sheets_changed": sorted({c["sheet"] for c in changed_cells}),
        "changed_cell_count": len(changed_cells),
        "changed_cells": changed_cells[:PREVIEW_MAX_CHANGED_CELLS],
        "notes": notes,
    }


class JobService:
    def __init__(
        self, root_dir: Path, llm: LLMClient | None = None, mode: str | None = None
    ) -> None:
        self.root_dir = root_dir
        self.root_dir.mkdir(parents=True, exist_ok=True)
        self.llm = llm or LLMClient()
        self.mode = (mode or AGENT_DEFAULT_MODE).lower()
        self.checker = CodeChecker()
        self.jobs: dict[str, Job] = {}
        self.download_tokens: dict[str, Path] = {}
        self.lock = threading.Lock()
        self.q: queue.Queue[str] = queue.Queue()
        self.worker = threading.Thread(target=self._worker_loop, daemon=True)
        self.worker.start()
        self.cleaner = threading.Thread(target=self._cleanup_loop, daemon=True)
        self.cleaner.start()

    def create_job(
        self, upload: UploadFile, instruction: str, sheet_name: str | None
    ) -> str:
        ext = Path(upload.filename or "").suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise JobError("UNSUPPORTED_FORMAT", "対応形式は .xlsx / .xlsm のみです")
        payload = upload.file.read(MAX_FILE_SIZE_BYTES + 1)
        if len(payload) > MAX_FILE_SIZE_BYTES:
            raise JobError("FILE_TOO_LARGE", "ファイルサイズが上限を超えています")
        try:
            uploaded_wb = load_workbook(
                io.BytesIO(payload), keep_vba=(ext == ".xlsm"), data_only=False
            )
            _close_workbook(uploaded_wb)
        except Exception as exc:
            raise JobError(
                "UNSUPPORTED_FORMAT", f"Excelファイルを開けませんでした: {exc}"
            ) from exc

        job_id = uuid.uuid4().hex
        work_dir = self.root_dir / job_id
        work_dir.mkdir(parents=True, exist_ok=True)
        source_path = work_dir / f"input{ext}"
        source_path.write_bytes(payload)

        with self.lock:
            self.jobs[job_id] = Job(
                job_id=job_id,
                instruction=instruction,
                sheet_name=sheet_name,
                created_at=time.time(),
                work_dir=work_dir,
                source_path=source_path,
            )
        self.q.put(job_id)
        return job_id

    def get_job(self, job_id: str) -> Job:
        with self.lock:
            job = self.jobs.get(job_id)
        if not job:
            raise KeyError(job_id)
        return job

    def approve_job(self, job_id: str) -> str:
        with self.lock:
            job = self.jobs.get(job_id)
            if not job:
                raise KeyError(job_id)
            if job.status != JobStatus.PREVIEW_READY or not job.result_path:
                raise JobError("INVALID_STATE", "プレビュー準備完了後のみ承認できます")
            job.status = JobStatus.APPROVED
            token = secrets.token_urlsafe(24)
            self.download_tokens[token] = job.result_path
            job.download_token = token
            job.status = JobStatus.DONE
            return token

    def pop_download_path(self, token: str) -> Path:
        with self.lock:
            path = self.download_tokens.pop(token, None)
        if path is None or not path.exists():
            raise KeyError(token)
        return path

    def _worker_loop(self) -> None:
        while True:
            job_id = self.q.get()
            try:
                self._process(job_id)
            finally:
                self.q.task_done()

    def _run_oneshot(
        self,
        job: Job,
        summary: dict[str, Any],
        blocked_cells: set[str],
        sheet_name: str,
        result_path: Path,
    ) -> None:
        """単発モード: コードを1回生成→ASTチェック→サンドボックス実行。"""
        feedback = ""
        validation_errors: list[str] = []
        for _ in range(LLM_MAX_RETRY):
            job.status = JobStatus.GENERATING
            generated_code = self.llm.generate_code(
                summary, job.instruction, feedback=feedback
            )
            job.status = JobStatus.CHECKING
            try:
                self.checker.validate(generated_code, blocked_cells)
                break
            except JobError as err:
                feedback = f"{err.error_code}: {err.message}"
                validation_errors.append(feedback)
        else:
            reason = (
                "; ".join(validation_errors)
                if validation_errors
                else "コード生成に失敗しました（検証詳細なし）"
            )
            raise JobError(
                "CODE_CHECK_FAILED", f"コード再生成上限に達しました: {reason}"
            )

        job.status = JobStatus.EXECUTING
        _exec_in_sandbox(result_path, sheet_name, generated_code)

    def _process(self, job_id: str) -> None:
        job = self.get_job(job_id)
        try:
            job.status = JobStatus.ANALYZING
            wb = load_workbook(
                job.source_path, keep_vba=_keep_vba_for(job.source_path), data_only=False
            )
            ws = (
                wb[job.sheet_name]
                if job.sheet_name and job.sheet_name in wb.sheetnames
                else wb.active
            )
            sheet_name = ws.title
            summary = _summarize_sheet(ws)
            blocked_cells = _non_anchor_cells(ws)
            _close_workbook(wb)

            result_path = job.work_dir / f"result{job.source_path.suffix}"
            shutil.copy2(job.source_path, result_path)

            if self.mode == "oneshot":
                self._run_oneshot(job, summary, blocked_cells, sheet_name, result_path)
            else:
                from app import agent

                job.status = JobStatus.GENERATING
                agent.run_agent(
                    result_path,
                    sheet_name,
                    job.instruction,
                    summary,
                    self.llm,
                    max_steps=AGENT_MAX_STEPS,
                    step_timeout=AGENT_STEP_TIMEOUT,
                )

            job.preview = _create_preview(job.source_path, result_path, sheet_name)
            job.result_path = result_path
            job.status = JobStatus.PREVIEW_READY
        except JobError as err:
            job.status = JobStatus.ERROR
            job.error_code = err.error_code
            job.message = err.message
            job.retryable = err.retryable
        except Exception as err:  # pragma: no cover
            job.status = JobStatus.ERROR
            job.error_code = "INTERNAL_ERROR"
            job.message = str(err)
            job.retryable = False

    def _cleanup_loop(self) -> None:
        while True:
            time.sleep(60)
            now = time.time()
            with self.lock:
                expired = [
                    job_id
                    for job_id, job in self.jobs.items()
                    if now - job.created_at > JOB_TTL_SECONDS
                ]
            for job_id in expired:
                with self.lock:
                    job = self.jobs.pop(job_id, None)
                    if job and job.download_token:
                        self.download_tokens.pop(job.download_token, None)
                if job and job.work_dir.exists():
                    shutil.rmtree(job.work_dir, ignore_errors=True)


CHAT_MAX_HISTORY = 6  # プロンプトに含める直近の依頼数
CHAT_UNDO_LIMIT = 20  # 保持する版数の上限


@dataclass
class ChatSession:
    session_id: str
    work_dir: Path
    ext: str
    sheet_name: str | None
    versions: list[Path]  # versions[0]=元ファイル, 末尾=現在の状態
    messages: list[dict[str, Any]] = field(default_factory=list)
    created_at: float = 0.0
    # 版ファイル名の一意採番（list長だとUndo上限トリム後に番号が衝突する）
    version_counter: int = 0
    # 同一セッションへの同時操作（メッセージ送信/Undo）を直列化する
    lock: threading.Lock = field(default_factory=threading.Lock)


class SessionService:
    """チャット形式の編集セッション。1ファイルをアップロード後、複数の指示を
    現在の状態に積み重ねて適用し、Undo で1手戻せる。"""

    def __init__(
        self, root_dir: Path, llm: LLMClient | None = None, mode: str | None = None
    ) -> None:
        self.root_dir = root_dir
        self.root_dir.mkdir(parents=True, exist_ok=True)
        self.llm = llm or LLMClient()
        self.mode = (mode or AGENT_DEFAULT_MODE).lower()
        self.sessions: dict[str, ChatSession] = {}
        self.lock = threading.Lock()
        self.cleaner = threading.Thread(target=self._cleanup_loop, daemon=True)
        self.cleaner.start()

    def create_session(self, upload: UploadFile, sheet_name: str | None) -> dict[str, Any]:
        ext = Path(upload.filename or "").suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise JobError("UNSUPPORTED_FORMAT", "対応形式は .xlsx / .xlsm のみです")
        payload = upload.file.read(MAX_FILE_SIZE_BYTES + 1)
        if len(payload) > MAX_FILE_SIZE_BYTES:
            raise JobError("FILE_TOO_LARGE", "ファイルサイズが上限を超えています")
        try:
            wb = load_workbook(io.BytesIO(payload), keep_vba=(ext == ".xlsm"), data_only=False)
            sheets = list(wb.sheetnames)
            active = (
                sheet_name if sheet_name and sheet_name in sheets else wb.active.title
            )
            _close_workbook(wb)
        except Exception as exc:
            raise JobError(
                "UNSUPPORTED_FORMAT", f"Excelファイルを開けませんでした: {exc}"
            ) from exc

        session_id = uuid.uuid4().hex
        work_dir = self.root_dir / session_id
        work_dir.mkdir(parents=True, exist_ok=True)
        v0 = work_dir / f"v0{ext}"
        v0.write_bytes(payload)
        with self.lock:
            self.sessions[session_id] = ChatSession(
                session_id=session_id,
                work_dir=work_dir,
                ext=ext,
                sheet_name=active,
                versions=[v0],
                created_at=time.time(),
            )
        return {"session_id": session_id, "sheets": sheets, "active_sheet": active}

    def _get(self, session_id: str) -> ChatSession:
        with self.lock:
            session = self.sessions.get(session_id)
        if not session:
            raise KeyError(session_id)
        return session

    def post_message(
        self, session_id: str, instruction: str, think: bool | None = None
    ) -> dict[str, Any]:
        if not instruction.strip():
            raise JobError("INVALID_STATE", "指示が空です")
        if self.mode == "oneshot":
            raise JobError("INVALID_STATE", "チャットは agent モードのみ対応です")
        session = self._get(session_id)
        # 同一セッションへの同時操作を直列化（version番号衝突や履歴破損を防ぐ）
        with session.lock:
            session.created_at = time.time()  # 利用中はTTLを延長
            current = session.versions[-1]

            wb = load_workbook(current, keep_vba=_keep_vba_for(current), data_only=False)
            ws = (
                wb[session.sheet_name]
                if session.sheet_name and session.sheet_name in wb.sheetnames
                else wb.active
            )
            sheet_name = ws.title
            summary = _summarize_sheet(ws)
            _close_workbook(wb)

            session.version_counter += 1
            candidate = session.work_dir / f"v{session.version_counter}{session.ext}"
            shutil.copy2(current, candidate)
            try:
                from app import agent

                agent.run_agent(
                    candidate,
                    sheet_name,
                    self._augment(session, instruction),
                    summary,
                    self.llm,
                    max_steps=AGENT_MAX_STEPS,
                    step_timeout=AGENT_STEP_TIMEOUT,
                    think=think,
                )
            except Exception:
                candidate.unlink(missing_ok=True)
                raise

            preview = _create_preview(current, candidate, sheet_name)
            session.versions.append(candidate)
            # Undo 上限を超えた古い版は削除（先頭の元ファイルは残す）
            while len(session.versions) > CHAT_UNDO_LIMIT + 1:
                old = session.versions.pop(1)
                old.unlink(missing_ok=True)
            session.messages.append({"role": "user", "text": instruction})
            session.messages.append({"role": "assistant", "preview": preview})
            can_undo = len(session.versions) > 1
        return {"preview": preview, "can_undo": can_undo}

    def undo(self, session_id: str) -> dict[str, Any]:
        session = self._get(session_id)
        with session.lock:
            session.created_at = time.time()
            if len(session.versions) <= 1:
                raise JobError("INVALID_STATE", "これ以上戻せません")
            removed = session.versions.pop()
            removed.unlink(missing_ok=True)
            # 取り消した手の user/assistant メッセージも履歴から除く
            # （残すと _augment が取り消し済みの指示を文脈に含めてしまう）
            if session.messages and session.messages[-1].get("role") == "assistant":
                session.messages.pop()
            if session.messages and session.messages[-1].get("role") == "user":
                session.messages.pop()
            can_undo = len(session.versions) > 1
        return {"ok": True, "can_undo": can_undo}

    def current_path(self, session_id: str) -> Path:
        session = self._get(session_id)
        return session.versions[-1]

    def _augment(self, session: ChatSession, instruction: str) -> str:
        """直近の依頼履歴を文脈として今回の指示に付与する（代名詞解決のため）。"""
        prior = [m["text"] for m in session.messages if m.get("role") == "user"]
        prior = prior[-CHAT_MAX_HISTORY:]
        if not prior:
            return instruction
        history = "\n".join(f"- {text}" for text in prior)
        return f"これまでの依頼:\n{history}\n\n今回の依頼: {instruction}"

    def _cleanup_loop(self) -> None:
        while True:
            time.sleep(60)
            now = time.time()
            with self.lock:
                expired = [
                    self.sessions.pop(sid)
                    for sid, s in list(self.sessions.items())
                    if now - s.created_at > JOB_TTL_SECONDS
                ]
            # 実行中のセッションを消さないよう、各セッションのロックを取ってから削除
            for session in expired:
                with session.lock:
                    if session.work_dir.exists():
                        shutil.rmtree(session.work_dir, ignore_errors=True)


def _to_job_response(job: Job) -> dict[str, Any]:
    data = {
        "job_id": job.job_id,
        "status": job.status,
    }
    if job.status == JobStatus.ERROR:
        data.update(
            {
                "error_code": job.error_code,
                "message": job.message,
                "retryable": job.retryable,
            }
        )
    if job.preview:
        data["preview"] = job.preview
    return data


def create_app(job_service: JobService | None = None) -> FastAPI:
    app = FastAPI(title="xlsx-agent", version="0.1")
    service = job_service or JobService(Path(os.getenv("JOB_ROOT", "./data/jobs")))
    chat = SessionService(
        Path(os.getenv("SESSION_ROOT", "./data/sessions")), llm=service.llm
    )

    # ブラウザ（別PC）からのアクセスを許可する。
    # 既定は全許可。特定のオリジンに限定したい場合は CORS_ORIGINS にカンマ区切りで指定。
    cors_origins = [
        origin.strip()
        for origin in os.getenv("CORS_ORIGINS", "*").split(",")
        if origin.strip()
    ]
    app.add_middleware(
        CORSMiddleware,
        allow_origins=cors_origins or ["*"],
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.exception_handler(Exception)
    async def _unhandled_exception(_request, exc: Exception) -> JSONResponse:
        # 未処理例外でも JSON で返す（フロントが "Internal Server Error" を
        # JSON.parse して落ちるのを防ぎ、原因をUIに表示できるようにする）。
        return JSONResponse(
            status_code=500,
            content={
                "error_code": "INTERNAL_ERROR",
                "message": f"{type(exc).__name__}: {exc}",
            },
        )

    @app.get("/healthz", include_in_schema=False)
    async def healthz() -> dict[str, Any]:
        # Ollama 疎通とモデルの存在も返す（メッセージ送信の400切り分けに使える）。
        # diagnose を持たない LLM が注入されても落ちないようにする。
        llm = service.llm
        info: dict[str, Any] = {
            "status": "ok",
            "model": getattr(llm, "model", None),
            "think": getattr(llm, "think", None),
        }
        diagnose = getattr(llm, "diagnose", None)
        if callable(diagnose):
            try:
                info.update(await run_in_threadpool(diagnose))
            except Exception:
                pass
        return info

    @app.get("/", include_in_schema=False)
    async def index() -> FileResponse:
        return FileResponse(STATIC_DIR / "chat.html")

    @app.get("/classic", include_in_schema=False)
    async def classic() -> FileResponse:
        return FileResponse(STATIC_DIR / "index.html")

    @app.get("/favicon.ico", include_in_schema=False)
    async def favicon() -> Response:
        # ブラウザが自動取得する favicon。アイコンは無いので 204 で静かに返す。
        return Response(status_code=204)

    # ---- チャット（セッション）API ----
    def _chat_error(err: JobError) -> HTTPException:
        return HTTPException(
            status_code=400,
            detail={
                "error_code": err.error_code,
                "message": err.message,
                "retryable": err.retryable,
            },
        )

    @app.post("/sessions")
    async def create_session(
        file: UploadFile = File(...),
        sheet_name: str | None = Form(default=None),
    ) -> dict[str, Any]:
        try:
            # 同期I/O(load_workbook等)でイベントループを塞がないよう別スレッドで実行
            return await run_in_threadpool(chat.create_session, file, sheet_name)
        except JobError as err:
            raise _chat_error(err) from err

    @app.post("/sessions/{session_id}/messages")
    async def post_message(
        session_id: str,
        instruction: str = Form(...),
        think: bool = Form(default=False),
    ) -> dict[str, Any]:
        try:
            return await run_in_threadpool(
                chat.post_message, session_id, instruction, think
            )
        except KeyError as err:
            raise HTTPException(status_code=404, detail="session not found") from err
        except JobError as err:
            raise _chat_error(err) from err

    @app.post("/sessions/{session_id}/undo")
    async def undo_session(session_id: str) -> dict[str, Any]:
        try:
            return chat.undo(session_id)
        except KeyError as err:
            raise HTTPException(status_code=404, detail="session not found") from err
        except JobError as err:
            raise _chat_error(err) from err

    @app.get("/sessions/{session_id}/download")
    async def download_session(session_id: str):
        try:
            path = chat.current_path(session_id)
        except KeyError as err:
            raise HTTPException(status_code=404, detail="session not found") from err
        return FileResponse(
            path,
            filename=f"edited{path.suffix}",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/jobs")
    async def create_job(
        file: UploadFile = File(...),
        instruction: str = Form(...),
        sheet_name: str | None = Form(default=None),
    ) -> dict[str, str]:
        try:
            job_id = service.create_job(file, instruction, sheet_name)
            return {"job_id": job_id, "status": JobStatus.QUEUED}
        except JobError as err:
            raise HTTPException(
                status_code=400,
                detail={
                    "status": JobStatus.ERROR,
                    "error_code": err.error_code,
                    "message": err.message,
                    "retryable": err.retryable,
                },
            ) from err

    @app.get("/jobs/{job_id}")
    async def get_job(job_id: str) -> dict[str, Any]:
        try:
            return _to_job_response(service.get_job(job_id))
        except KeyError as err:
            raise HTTPException(status_code=404, detail="job not found") from err

    @app.post("/jobs/{job_id}/approve")
    async def approve(job_id: str) -> dict[str, str]:
        try:
            token = service.approve_job(job_id)
            return {"job_id": job_id, "download_url": f"/download/{token}"}
        except KeyError as err:
            raise HTTPException(status_code=404, detail="job not found") from err
        except JobError as err:
            raise HTTPException(status_code=400, detail=err.message) from err

    @app.get("/download/{token}")
    async def download(token: str):
        try:
            path = service.pop_download_path(token)
            return FileResponse(
                path,
                filename=f"edited{path.suffix}",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except KeyError as err:
            raise HTTPException(status_code=404, detail="token not found") from err

    if STATIC_DIR.is_dir():
        app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

    return app


app = create_app()
