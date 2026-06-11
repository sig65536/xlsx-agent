import io
import multiprocessing as mp
import os
import sys
import time
from pathlib import Path

import pytest
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook

from app.agent import (
    _is_allowed_import,
    parse_action,
    precheck_step_code,
    run_agent,
)
from app.main import JobError, JobService


def _hardening_probe(conn):
    """spawn子プロセスで hardening を適用し、結果を親へ返す（テスト用）。"""
    import resource

    from app.agent import _harden_worker_process

    _harden_worker_process()
    soft, _hard = resource.getrlimit(resource.RLIMIT_NOFILE)
    conn.send(
        {
            "nofile": soft,
            "config_scrubbed": "XLSX_AGENT_WORKER_NOFILE" not in os.environ,
        }
    )


def _network_probe(conn):
    """spawn子プロセスでネットワーク遮断を適用し、結果を親へ返す（テスト用）。"""
    import socket

    from app.agent import _disable_network

    _disable_network()
    result = {"is_class": isinstance(socket.socket, type)}
    try:
        socket.socket()
        result["socket"] = "open"
    except OSError:
        result["socket"] = "blocked"
    try:
        socket.getaddrinfo("example.com", 80)
        result["dns"] = "open"
    except OSError:
        result["dns"] = "blocked"
    try:
        socket.gethostbyname("example.com")
        result["legacy_dns"] = "open"
    except OSError:
        result["legacy_dns"] = "blocked"
    conn.send(result)


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "before"
    ws["B1"] = 10
    ws["B2"] = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class StepLLM:
    """与えられたコード列を順に返し、最後に DONE を返すスタブ。"""

    def __init__(self, codes: list[str]):
        self._codes = codes
        self.calls = 0

    def agent_step(self, summary, instruction, transcript):
        idx = self.calls
        self.calls += 1
        if idx < len(self._codes):
            return f"```python\n{self._codes[idx]}\n```"
        return "DONE"


def _run_agent_on(tmp_path: Path, codes: list[str]) -> Path:
    src = tmp_path / "input.xlsx"
    src.write_bytes(_workbook_bytes())
    summary = {"sheet_name": "Sheet"}
    run_agent(src, "Sheet", "テスト", summary, StepLLM(codes), max_steps=6, step_timeout=30)
    return src


def test_parse_action_code_and_done() -> None:
    kind, code = parse_action("```python\nws['A1']=1\n```")
    assert kind == "code" and code == "ws['A1']=1"
    assert parse_action("DONE")[0] == "done"
    assert parse_action("All set.\nDONE.")[0] == "done"
    # 散文や "not done" は完了扱いにしない（部分保存を防ぐ）
    assert parse_action("特に何もありません")[0] == "malformed"
    assert parse_action("I am not done yet")[0] == "malformed"


def test_precheck_rejects_escape() -> None:
    with pytest.raises(JobError):
        precheck_step_code("x = ws.__class__")
    with pytest.raises(JobError):
        precheck_step_code("open('/etc/passwd')")
    # 許可モジュール経由の属性チェーン脱獄を遮断（os/sys/system を属性として禁止）
    with pytest.raises(JobError):
        precheck_step_code("import random\nrandom.os.system('whoami')")
    with pytest.raises(JobError):
        precheck_step_code("import os\nos.listdir('/')")
    with pytest.raises(JobError):
        precheck_step_code("x.subprocess.Popen(['ls'])")
    # プライベート別名(_os/_sys)経由・__builtins__ 参照も遮断
    with pytest.raises(JobError):
        precheck_step_code("import random\nrandom._os.system('whoami')")
    with pytest.raises(JobError):
        precheck_step_code("re._sys.modules")
    with pytest.raises(JobError):
        precheck_step_code("print(__builtins__)")
    # import の別名経由（ast.alias）の禁止名/ダンダーを遮断
    with pytest.raises(JobError):
        precheck_step_code("import os as x\nx.listdir('/')")
    with pytest.raises(JobError):
        precheck_step_code("from datetime import __class__ as c")
    # 文字列フォーマット経由の属性チェーン脱獄を遮断（ダンダー/単一アンダースコア両方）
    with pytest.raises(JobError):
        precheck_step_code("ws['A1'] = '{0.__class__.__init__}'.format(ws)")
    with pytest.raises(JobError):
        precheck_step_code("ws['A1'] = '{0._os.environ}'.format(random)")


def test_safe_import_blocks_dangerous_fromlist_alias() -> None:
    from app.agent import _safe_import

    # `from random import _os as x` 型の別名脱獄を import 時点で遮断
    with pytest.raises(ImportError):
        _safe_import("random", fromlist=("_os",))
    with pytest.raises(ImportError):
        _safe_import("datetime", fromlist=("os",))
    # 正当な書式系 import（from-import）は通る
    assert _safe_import("openpyxl.styles", fromlist=("Font",)) is not None
    # plain `import openpyxl.utils` は親 openpyxl(load_workbook) を束縛するため拒否
    with pytest.raises(ImportError):
        _safe_import("openpyxl.utils", fromlist=())
    assert _safe_import("openpyxl.utils", fromlist=("get_column_letter",)) is not None


def test_precheck_allows_re_compile() -> None:
    # 属性 .compile（re.compile 等）は誤検知で弾かない
    precheck_step_code("import re\np = re.compile('[0-9]+')")


def test_safe_builtins_is_minimal_allowlist() -> None:
    from app.agent import _make_safe_builtins

    builtins = _make_safe_builtins()
    # site が注入する開示系・危険系は含まれない
    for name in ("license", "copyright", "credits", "help", "open", "eval",
                 "exec", "getattr", "globals", "input", "exit", "quit"):
        assert name not in builtins, name
    # Excel編集に必要な基本ビルトインは含まれる
    for name in ("range", "len", "print", "sorted", "enumerate", "str", "int"):
        assert name in builtins, name


def test_import_whitelist_blocks_openpyxl_root_and_io() -> None:
    # 書式系サブモジュールは許可
    assert _is_allowed_import("openpyxl.styles") is True
    assert _is_allowed_import("openpyxl.utils") is True
    assert _is_allowed_import("datetime") is True
    # openpyxl 直下（load_workbook 等のI/O経路）と危険モジュールは不許可
    assert _is_allowed_import("openpyxl") is False
    assert _is_allowed_import("openpyxl.reader.excel") is False
    assert _is_allowed_import("os") is False
    assert _is_allowed_import("subprocess") is False


def test_agent_rolls_back_failed_step(tmp_path: Path) -> None:
    """失敗ステップの途中変更は破棄され、保存に残らない。"""
    out = _run_agent_on(
        tmp_path,
        [
            "ws['A1'] = 'partial'\n1 / 0",   # A1 を書いた直後に例外 → ロールバック
            "ws['B1'] = 'good'",             # こちらは成功 → 保持
        ],
    )
    wb = load_workbook(out)
    assert wb.active["A1"].value == "before"  # 失敗ステップの変更は残らない
    assert wb.active["B1"].value == "good"


def test_agent_preserves_vars_across_rollback(tmp_path: Path) -> None:
    """失敗ステップのロールバック後も、成功ステップで定義した変数は残る。"""
    out = _run_agent_on(
        tmp_path,
        [
            "counter = 5",                       # 変数定義（成功）
            "ws['A1'] = str(counter)\n1 / 0",    # 途中で失敗 → ロールバック
            "ws['B1'] = str(counter)",           # counter が残っていれば成功
        ],
    )
    wb = load_workbook(out)
    assert wb.active["A1"].value == "before"  # 失敗ステップは巻き戻る
    assert wb.active["B1"].value == "5"        # counter は保持されている


def test_agent_step_limit_does_not_save_partial(tmp_path: Path) -> None:
    """DONEに到達せずステップ上限で打ち切ったら部分保存せずエラーにする。"""

    class NeverDoneLLM:
        def agent_step(self, summary, instruction, transcript):
            return "```python\nws['A1'] = 'x'\n```"

    src = tmp_path / "input.xlsx"
    src.write_bytes(_workbook_bytes())
    with pytest.raises(JobError) as err:
        run_agent(
            src, "Sheet", "テスト", {"sheet_name": "Sheet"}, NeverDoneLLM(), max_steps=2
        )
    assert err.value.error_code == "AGENT_STEP_LIMIT"


def test_agent_multistep_edit(tmp_path: Path) -> None:
    out = _run_agent_on(
        tmp_path,
        [
            "ws['A1'] = 'edited'",
            "ws['B3'] = ws['B1'].value + ws['B2'].value",
        ],
    )
    wb = load_workbook(out)
    assert wb.active["A1"].value == "edited"
    assert wb.active["B3"].value == 30


def test_agent_allows_whitelisted_import_for_formatting(tmp_path: Path) -> None:
    """openpyxl.styles の import が通り、書式設定ができること（緩和の要点）。"""
    out = _run_agent_on(
        tmp_path,
        [
            "from openpyxl.styles import Font\nws['A1'].font = Font(bold=True)",
        ],
    )
    wb = load_workbook(out)
    assert wb.active["A1"].font.bold is True


def test_agent_blocks_dangerous_import(tmp_path: Path) -> None:
    """os の import はサンドボックスで失敗し、ジョブ全体は失敗扱いになること。"""
    with pytest.raises(JobError):
        _run_agent_on(tmp_path, ["import os\nos.listdir('/')"])


def test_preview_detects_cross_sheet_edits(tmp_path: Path) -> None:
    """他シートへの編集もプレビューに出る（見ていない変更を承認させない）。"""
    from app.main import _create_preview

    wb = Workbook()
    wb.active["A1"] = "x"
    wb.create_sheet("Other")["A1"] = "old"
    before = tmp_path / "b.xlsx"
    wb.save(before)
    wb["Other"]["A1"] = "new"  # アクティブでないシートを編集
    after = tmp_path / "a.xlsx"
    wb.save(after)

    preview = _create_preview(before, after, wb.active.title)
    assert preview["changed_cell_count"] >= 1
    assert "Other" in preview["sheets_changed"]


def test_scrub_env_keeps_system_drops_secrets() -> None:
    from app.agent import _scrub_env

    env = {
        "PATH": "/usr/bin",
        "SYSTEMROOT": "C:/Windows",
        "OLLAMA_ENDPOINT": "http://x",
        "GITHUB_TOKEN": "secret",
        "MY_API_KEY": "secret",
    }
    _scrub_env(env)
    assert env.get("PATH") == "/usr/bin"
    assert "SYSTEMROOT" in env
    assert "GITHUB_TOKEN" not in env
    assert "MY_API_KEY" not in env
    assert "OLLAMA_ENDPOINT" not in env


@pytest.mark.skipif(sys.platform == "win32", reason="resource limits are POSIX-only")
def test_worker_reads_config_before_env_scrub() -> None:
    """XLSX_AGENT_WORKER_* はスクラブより先に読まれる（順序バグの回帰テスト）。"""
    ctx = mp.get_context("spawn")
    parent, child = ctx.Pipe()
    os.environ["XLSX_AGENT_WORKER_NOFILE"] = "128"
    os.environ["XLSX_AGENT_DISABLE_NETWORK"] = "0"
    try:
        proc = ctx.Process(target=_hardening_probe, args=(child,))
        proc.start()
        result = parent.recv()
        proc.join(10)
    finally:
        os.environ.pop("XLSX_AGENT_WORKER_NOFILE", None)
        os.environ.pop("XLSX_AGENT_DISABLE_NETWORK", None)
    assert result["nofile"] == 128  # スクラブ前に設定を読めている
    assert result["config_scrubbed"] is True  # その後 env はスクラブされている


def test_disable_network_blocks_sockets() -> None:
    # ネットワーク遮断は socket.socket.__init__ をパッチするため、メインの
    # テストプロセスを汚さないよう spawn 子プロセスで検証する。
    ctx = mp.get_context("spawn")
    parent, child = ctx.Pipe()
    proc = ctx.Process(target=_network_probe, args=(child,))
    proc.start()
    result = parent.recv()
    proc.join(10)
    assert result["socket"] == "blocked"
    assert result["dns"] == "blocked"           # DNS(getaddrinfo)も遮断
    assert result["legacy_dns"] == "blocked"     # 旧resolver(gethostbyname)も遮断
    assert result["is_class"] is True            # isinstance 互換が壊れていない


def test_jobservice_agent_mode_lifecycle(tmp_path: Path) -> None:
    class AgentStub:
        def agent_step(self, summary, instruction, transcript):
            if not transcript:
                return "```python\nws['A1'] = 'edited'\n```"
            return "DONE"

    service = JobService(tmp_path / "jobs", llm=AgentStub(), mode="agent")
    upload = UploadFile(file=io.BytesIO(_workbook_bytes()), filename="sample.xlsx")
    job_id = service.create_job(upload, "A1をeditedに", None)
    for _ in range(60):
        job = service.get_job(job_id)
        if job.status in ("preview_ready", "error"):
            break
        time.sleep(0.1)
    assert job.status == "preview_ready", f"{job.error_code}: {job.message}"
    assert job.preview["changed_cell_count"] >= 1
