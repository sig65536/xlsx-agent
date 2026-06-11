import asyncio
import io
import time
from pathlib import Path

import pytest
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook

from app.main import JobService, create_app


class StubLLM:
    def generate_code(self, summary, instruction, feedback=""):
        return "ws['A1'] = 'edited'"


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "before"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(filename: str) -> UploadFile:
    return UploadFile(file=io.BytesIO(_workbook_bytes()), filename=filename)


def _download_endpoint(app):
    for route in app.routes:
        if getattr(route, "path", None) == "/download/{token}":
            return route.endpoint
    raise AssertionError("download endpoint not found")


def _run_lifecycle(tmp_path: Path, filename: str) -> tuple[JobService, str, bytes, str]:
    service = JobService(tmp_path / f"jobs-{Path(filename).suffix[1:]}", llm=StubLLM())
    job_id = service.create_job(_upload(filename), "A1をeditedに変更", None)

    status = "queued"
    job = service.get_job(job_id)
    for _ in range(40):
        job = service.get_job(job_id)
        status = job.status
        if status == "preview_ready":
            break
        time.sleep(0.1)
    assert status == "preview_ready"
    assert job.preview is not None
    assert job.preview["changed_cell_count"] >= 1

    token = service.approve_job(job_id)
    app = create_app(service)
    response = asyncio.run(_download_endpoint(app)(token))
    assert response.status_code == 200
    with open(response.path, "rb") as result_file:
        content = result_file.read()
    out = load_workbook(io.BytesIO(content), data_only=False)
    assert out.active["A1"].value == "edited"
    out.close()

    try:
        service.pop_download_path(token)
        assert False, "Expected one-time download token to be consumed"
    except KeyError:
        pass
    return service, job_id, content, response.headers["content-disposition"]


def test_job_lifecycle(tmp_path: Path) -> None:
    service, job_id, _content, content_disposition = _run_lifecycle(
        tmp_path, "sample.xlsx"
    )

    job = service.get_job(job_id)
    assert job.source_path.suffix == ".xlsx"
    assert job.result_path is not None
    assert job.result_path.name == "result.xlsx"
    assert 'filename="edited.xlsx"' in content_disposition


def test_xlsm_job_lifecycle_preserves_result_and_download_extensions(
    tmp_path: Path,
) -> None:
    service, job_id, _content, content_disposition = _run_lifecycle(
        tmp_path, "sample.xlsm"
    )

    job = service.get_job(job_id)
    assert job.source_path.suffix == ".xlsm"
    assert job.result_path is not None
    assert job.result_path.name == "result.xlsm"
    assert 'filename="edited.xlsm"' in content_disposition


def test_keep_vba_policy_follows_extension() -> None:
    from app.main import _keep_vba_for

    assert _keep_vba_for("foo.xlsm") is True
    assert _keep_vba_for("foo.xlsx") is False
    assert _keep_vba_for(Path("dir/result.xlsx")) is False


def test_xlsx_result_is_not_macro_enabled(tmp_path: Path) -> None:
    """回帰テスト: .xlsx を往復しても macroEnabled content-type にならないこと。

    keep_vba=True を .xlsx に使うと content-type が macroEnabled になり、
    openpyxl では開けるが Excel が「破損」と判断する。これを防ぐ。
    """
    import zipfile

    _service, _job_id, content, _cd = _run_lifecycle(tmp_path, "sample.xlsx")
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        content_types = zf.read("[Content_Types].xml").decode()
    assert "macroEnabled" not in content_types


def test_validate_excel_file_rejects_macro_content_type_on_xlsx(tmp_path: Path) -> None:
    from app.main import JobError, validate_excel_file

    bad = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    # わざと keep_vba=True 相当の macroEnabled な状態で保存して壊れを再現
    wb.save(bad)
    # content-type を macroEnabled に差し替えて Excel 破損状態を作る
    import shutil
    import zipfile

    tampered = tmp_path / "tampered.xlsx"
    with zipfile.ZipFile(bad) as zin, zipfile.ZipFile(tampered, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = data.replace(
                    b"application/vnd.openxmlformats-officedocument."
                    b"spreadsheetml.sheet.main+xml",
                    b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
            zout.writestr(item, data)
    shutil.copy2(tampered, tmp_path / "final.xlsx")
    with pytest.raises(JobError) as err:
        validate_excel_file(tmp_path / "final.xlsx")
    assert err.value.error_code == "EXCEL_SAVE_VALIDATION_FAILED"
