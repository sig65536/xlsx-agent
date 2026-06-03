from openpyxl import Workbook

from app.main import _summarize_sheet


def test_summarize_sheet_includes_structure_details() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:C1")
    ws["A1"] = "売上一覧"
    ws["A2"] = "日付"
    ws["B2"] = "商品名"
    ws["C2"] = "売上"
    ws["A3"] = "2026-06-03"
    ws["B3"] = "ノート"
    ws["C3"] = 100
    ws["C4"] = "=SUM(C3:C3)"
    ws.row_dimensions[5].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws.protection.sheet = True

    summary = _summarize_sheet(ws)

    assert summary["sample_values"][0]["row"] == 1
    assert summary["sample_values"][0]["values"][:3] == ["売上一覧", None, None]
    assert len(summary["sample_values"][0]["values"]) == 20
    assert summary["hidden_rows"] == [5]
    assert summary["hidden_columns"] == ["D"]
    assert summary["protected"] is True
    assert summary["merged_cells"] == [
        {"range": "A1:C1", "anchor": "A1", "value": "売上一覧"}
    ]
    assert any(
        candidate["row"] == 2 and candidate["values"][:3] == ["日付", "商品名", "売上"]
        for candidate in summary["header_candidates"]
    )
    assert summary["table_like_ranges"][0]["range"] == "A2:C4"
    assert summary["formula_cells"] == ["C4"]
